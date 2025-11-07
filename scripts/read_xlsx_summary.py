from openpyxl import load_workbook
import sys
wb = load_workbook(sys.argv[1], read_only=True)
print('Sheets:', wb.sheetnames)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    rows = []
    for i,row in enumerate(ws.iter_rows(values_only=True), start=1):
        rows.append(row)
        if i>=5:
            break
    print('\nSheet:', sheet)
    for r in rows:
        print(r)
